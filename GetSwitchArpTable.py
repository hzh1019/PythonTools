import paramiko
import openpyxl

# create ssh client
ssh = paramiko.SSHClient()
ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

# connect to switch
ssh.connect(hostname='192.168.1.1', port=22, username='admin', password='admin')

# execute command to get arp table
stdin, stdout, stderr = ssh.exec_command('display arp')

# read output and parse arp table
# This example applies only to the H3C S10000 series based on the processing results of different switches
arp_table = []
for line in stdout:
    if 'GE' in line:
        arp_entry = line.split()
        arp_table.append({'ip': arp_entry[0], 'mac': arp_entry[1], 'int': arp_entry[3]})

# create excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# write arp table to worksheet
worksheet['A1'] = 'IP Address'
worksheet['B1'] = 'MAC Address'
worksheet['C1'] = 'Interface'
for i, arp_entry in enumerate(arp_table):
    worksheet.cell(row=i+2, column=1, value=arp_entry['ip'])
    worksheet.cell(row=i+2, column=2, value=arp_entry['mac'])
    worksheet.cell(row=i+2, column=3, value=arp_entry['int'])

# save workbook to file
workbook.save('arp.xlsx')

# close ssh connection
ssh.close()

